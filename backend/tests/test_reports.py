"""Report generation, the schedule that drives it, and the drill-through that makes it useful.

The tests that matter most here are the ones that would fail if a shortcut had been taken: that a
generated file is a real PDF and a real XLSX rather than a placeholder, that its reference resolves
back to the query behind it, that two identical requests produce two rows rather than one being
overwritten, that a figure on a report carries a filter which genuinely reproduces it, and that a
failed AI call costs the report its paragraph and nothing else.
"""

from __future__ import annotations

import io
import zipfile
from datetime import timedelta, timezone

import pytest
from sqlalchemy import func, select

from app.core.roles import PlatformRole
from app.db.base import utcnow
from app.models.audit import AuditEvent
from app.models.enums import (
    ExceptionCategory,
    IntegrationJobStatus,
    IntegrationTargetSystem,
    TransactionStatus,
)
from app.models.reporting import Report
from app.models.transactions import TradeTransaction
from app.services.analytics import kpis, report_service, schedule
from app.services.analytics.report_templates import TEMPLATES, template_for
from app.services.gemini_service import AIServiceError
from tests.utils.analytics import (
    account,
    approve,
    integration_job,
    open_exception,
    transaction_at,
)

NOW = utcnow()


def utc(moment):
    """A stored moment, as an aware UTC one.

    PostgreSQL hands back what it was given; the SQLite fallback drops the offset on the way in
    and returns a naive value for the same column. Everything this platform stores is UTC, so
    re-attaching the zone is a rendering of the same instant rather than a reinterpretation - and
    it keeps a schedule assertion about *when* from turning into an assertion about which database
    the suite happened to run against.
    """
    return moment if moment.tzinfo is not None else moment.replace(tzinfo=timezone.utc)


def a_request(
    *,
    report_type: str = "adhoc",
    output_format: str = "pdf",
    days: int = 30,
    stream: str = "both",
    status: str | None = None,
) -> report_service.ReportRequest:
    return report_service.validate_request(
        report_type=report_type,
        output_format=output_format,
        period=kpis.Period(start=NOW - timedelta(days=days), end=NOW + timedelta(minutes=1)),
        stream=stream,
        status_filter=status,
    )


async def seeded_period(db_session):
    """A month of trading with enough variety that every figure is non-trivial."""
    approved = await transaction_at(
        db_session,
        batch_number="RPT-1",
        created_at=NOW - timedelta(days=6),
        request_created_at=NOW - timedelta(days=6),
        status=TransactionStatus.APPROVED.value,
    )
    await approve(db_session, approved, decided_at=NOW - timedelta(days=5))

    intervened = await transaction_at(
        db_session,
        batch_number="RPT-2",
        created_at=NOW - timedelta(days=4),
        request_created_at=NOW - timedelta(days=4),
        status=TransactionStatus.APPROVED.value,
    )
    await approve(db_session, intervened, decided_at=NOW - timedelta(days=2))
    await open_exception(
        db_session,
        exception_type=ExceptionCategory.QUANTITY_VARIATION_OUTSIDE_TOLERANCE.value,
        owner_role=PlatformRole.PURCHASE_USER.value,
        opened_at=NOW - timedelta(days=3),
        transaction=intervened,
        resolved_at=NOW - timedelta(days=2, hours=6),
    )

    open_case = await transaction_at(
        db_session, batch_number="RPT-3", created_at=NOW - timedelta(days=2)
    )
    await open_exception(
        db_session,
        exception_type=ExceptionCategory.LOW_CONFIDENCE.value,
        owner_role=PlatformRole.PURCHASE_USER.value,
        opened_at=NOW - timedelta(hours=90),
        transaction=open_case,
    )
    await integration_job(
        db_session,
        approved,
        target_system=IntegrationTargetSystem.SAP.value,
        status=IntegrationJobStatus.AWAITING_MANUAL_ACTION.value,
    )
    await integration_job(
        db_session,
        approved,
        target_system=IntegrationTargetSystem.DMS.value,
        status=IntegrationJobStatus.FAILED.value,
    )
    await db_session.commit()
    return approved, intervened, open_case


# --- real output ------------------------------------------------------------------------------------


async def test_pdf_generation_produces_a_real_pdf_carrying_its_reference(db_session, storage_root):
    await seeded_period(db_session)
    user = await account(db_session, roles=[PlatformRole.ADMIN.value], name="Aisha Rahman")
    await db_session.commit()

    report = await report_service.generate(db_session, a_request(), requested_by=user, now=NOW)
    await db_session.commit()

    assert report.output_format == "pdf"
    assert report.storage_ref.endswith(".pdf")

    data = await storage_root.download(report.storage_ref)
    assert data.startswith(b"%PDF")
    assert report.byte_size == len(data)

    import pymupdf

    with pymupdf.open(stream=data, filetype="pdf") as document:
        assert document.page_count >= 1
        text = "\n".join(page.get_text() for page in document)

    # The reference is printed in the document itself, which is what makes a page on somebody's
    # desk resolvable back to the query behind it.
    assert report.generation_reference in text
    assert "AGFZE Command Centre" in text
    # And the document says on its own face that nothing was sent anywhere.
    assert "never sent by the platform itself" in text


async def test_xlsx_generation_produces_a_real_workbook_with_the_reference_on_it(
    db_session, storage_root
):
    await seeded_period(db_session)
    user = await account(db_session, roles=[PlatformRole.APPROVER_HOD.value])
    await db_session.commit()

    report = await report_service.generate(
        db_session, a_request(output_format="xlsx"), requested_by=user, now=NOW
    )
    await db_session.commit()

    data = await storage_root.download(report.storage_ref)
    assert zipfile.is_zipfile(io.BytesIO(data))

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data))
    assert "Report" in workbook.sheetnames
    cover = "\n".join(
        str(cell.value)
        for row in workbook["Report"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert report.generation_reference in cover
    assert "this file is never sent by the platform" in cover.lower()
    # Every table section becomes its own sheet, so a recipient can sort and total a column.
    assert len(workbook.sheetnames) > 1


async def test_the_reference_resolves_back_to_the_query_and_the_audit_row(db_session, storage_root):
    await seeded_period(db_session)
    user = await account(db_session, roles=[PlatformRole.ADMIN.value])
    await db_session.commit()

    report = await report_service.generate(
        db_session,
        a_request(stream="scrap", status=TransactionStatus.APPROVED.value),
        requested_by=user,
        now=NOW,
    )
    await db_session.commit()

    resolved = await report_service.resolve_reference(db_session, report.generation_reference)
    assert resolved is not None
    assert resolved.id == report.id
    assert resolved.parameters["stream"] == "scrap"
    assert resolved.parameters["status_filter"] == TransactionStatus.APPROVED.value
    assert resolved.parameters["period_start"] == report.period_start.isoformat()

    event = await db_session.get(AuditEvent, report.audit_event_id)
    assert event is not None
    assert event.event_type == report_service.AuditEvent.REPORT_GENERATED
    assert event.event_metadata["generation_reference"] == report.generation_reference
    # The audit trail says as plainly as the document does that nothing was sent.
    assert event.event_metadata["distributed"] is False


async def test_every_reference_is_unique_across_generations(db_session, storage_root):
    await seeded_period(db_session)
    user = await account(db_session, roles=[PlatformRole.ADMIN.value])
    await db_session.commit()

    references = set()
    for _ in range(5):
        report = await report_service.generate(db_session, a_request(), requested_by=user, now=NOW)
        await db_session.commit()
        references.add(report.generation_reference)

    assert len(references) == 5


# --- versioning: never overwritten -------------------------------------------------------------------


async def test_two_identical_requests_produce_two_distinct_reports(db_session, storage_root):
    await seeded_period(db_session)
    user = await account(db_session, roles=[PlatformRole.ADMIN.value])
    await db_session.commit()

    request = a_request()
    first = await report_service.generate(db_session, request, requested_by=user, now=NOW)
    await db_session.commit()
    second = await report_service.generate(db_session, request, requested_by=user, now=NOW)
    await db_session.commit()

    assert first.id != second.id
    assert first.generation_reference != second.generation_reference
    assert first.storage_ref != second.storage_ref
    assert first.parameters == second.parameters

    total = await db_session.scalar(select(func.count()).select_from(Report))
    assert total == 2

    # Both files still exist. Neither generation overwrote the other's bytes.
    assert await storage_root.download(first.storage_ref)
    assert await storage_root.download(second.storage_ref)


# --- drill-through ------------------------------------------------------------------------------------


async def test_every_figure_on_a_report_carries_a_filter_that_reproduces_it(
    db_session, storage_root
):
    await seeded_period(db_session)
    user = await account(db_session, roles=[PlatformRole.ADMIN.value])
    await db_session.commit()

    report = await report_service.generate(db_session, a_request(), requested_by=user, now=NOW)
    await db_session.commit()

    navigable = 0
    for section in report.content["sections"]:
        for figure in section.get("figures", []):
            if figure.get("target"):
                navigable += 1
                assert isinstance(figure["filters"], dict)
        for row in section.get("rows", []):
            if row.get("target"):
                navigable += 1
                assert isinstance(row["filters"], dict)
    assert navigable > 0


async def test_a_status_figure_drills_through_to_exactly_the_rows_behind_it(
    db_session, storage_root
):
    """Follow the filter the report published and count what comes back."""
    await seeded_period(db_session)
    user = await account(db_session, roles=[PlatformRole.ADMIN.value])
    await db_session.commit()

    report = await report_service.generate(db_session, a_request(), requested_by=user, now=NOW)
    await db_session.commit()

    section = next(row for row in report.content["sections"] if row["key"] == "transactions")
    approved_row = next(
        row for row in section["rows"] if row["filters"].get("status") == "approved"
    )
    assert approved_row["count"] == 2

    # Now run the query that filter describes, exactly as the Transactions list would.
    filters = approved_row["filters"]
    statement = select(func.count()).select_from(
        select(TradeTransaction)
        .where(
            TradeTransaction.status == filters["status"],
            TradeTransaction.created_at >= report.period_start,
            TradeTransaction.created_at < report.period_end,
        )
        .subquery()
    )
    assert int(await db_session.scalar(statement)) == approved_row["count"]


async def test_an_exception_figure_drills_through_to_the_open_cases_it_counted(
    db_session, storage_root
):
    await seeded_period(db_session)
    user = await account(db_session, roles=[PlatformRole.ADMIN.value])
    await db_session.commit()

    report = await report_service.generate(db_session, a_request(), requested_by=user, now=NOW)
    await db_session.commit()

    section = next(row for row in report.content["sections"] if row["key"] == "exceptions")
    row = next(
        item
        for item in section["rows"]
        if item["filters"]["exception_type"] == ExceptionCategory.LOW_CONFIDENCE.value
    )
    assert row["open_count"] == 1
    assert row["filters"]["status"] == "open"
    assert row["target"] == "exceptions"

    from app.models.governance import ExceptionCase

    counted = await db_session.scalar(
        select(func.count(ExceptionCase.id)).where(
            ExceptionCase.exception_type == row["filters"]["exception_type"],
            ExceptionCase.resolved_at.is_(None),
        )
    )
    assert int(counted) == row["open_count"]


async def test_the_detail_table_links_each_row_to_its_own_batch(db_session, storage_root):
    await seeded_period(db_session)
    user = await account(db_session, roles=[PlatformRole.ADMIN.value])
    await db_session.commit()

    report = await report_service.generate(db_session, a_request(), requested_by=user, now=NOW)
    await db_session.commit()

    section = next(row for row in report.content["sections"] if row["key"] == "detail")
    assert section["rows"]
    for row in section["rows"]:
        assert row["transaction_id"]
        assert await db_session.get(
            TradeTransaction, __import__("uuid").UUID(row["transaction_id"])
        )


async def test_a_transaction_with_no_figures_on_it_still_renders_and_stores(
    db_session, storage_root
):
    """A gap has to survive the whole pipeline as a gap.

    pandas represents a missing number as NaN, JSON cannot hold one, and a deal genuinely can
    have no invoiced amount recorded against it yet. The absence has to travel through the frame,
    the stored content and the rendered document as an absence - not as a NaN, and certainly not
    as a zero somebody could read as a real figure.
    """
    from tests.utils.transactions import make_request, make_transaction

    request = await make_request(db_session)
    bare = await make_transaction(
        db_session,
        request=request,
        batch_number="RPT-BARE",
        quantity=None,
        rate=None,
        amount=None,
        supplier_name=None,
        contract_number=None,
        invoice_number=None,
    )
    bare.created_at = NOW - timedelta(days=1)
    user = await account(db_session, roles=[PlatformRole.ADMIN.value])
    await db_session.commit()

    report = await report_service.generate(db_session, a_request(), requested_by=user, now=NOW)
    await db_session.commit()

    detail = next(s for s in report.content["sections"] if s["key"] == "detail")
    row = next(r for r in detail["rows"] if r["batch_number"] == bare.batch_number)

    assert row["quantity_mt"] is None
    assert row["value"] is None
    assert row["counterparty"] is None

    # The stored content survives a round trip through the database's own JSON type, which a NaN
    # would not.
    refreshed = await report_service.get_report(db_session, report.id)
    assert refreshed.content["sections"]

    data = await storage_root.download(report.storage_ref)
    import pymupdf

    with pymupdf.open(stream=data, filetype="pdf") as document:
        text = "\n".join(page.get_text() for page in document)
    assert "nan" not in text.lower()
    assert bare.batch_number in text


# --- the AI paragraph ----------------------------------------------------------------------------------


async def test_a_failed_ai_summary_does_not_stop_a_monthly_report_generating(
    db_session, storage_root, monkeypatch
):
    """The failure that must cost nothing but the paragraph."""
    await seeded_period(db_session)
    user = await account(db_session, roles=[PlatformRole.APPROVER_HOD.value])
    await db_session.commit()

    async def refuse(_facts):
        raise AIServiceError(reason="quota_exhausted")

    monkeypatch.setattr(report_service, "summarize_reporting_period", refuse)

    report = await report_service.generate(
        db_session, a_request(report_type="monthly"), requested_by=user, now=NOW
    )
    await db_session.commit()

    assert report.ai_summary_error == "quota_exhausted"

    summary_section = next(
        section for section in report.content["sections"] if section["kind"] == "ai_summary"
    )
    assert summary_section["text"] is None
    assert summary_section["ai_generated"] is False
    assert "unaffected" in summary_section["unavailable_reason"]

    # Every deterministic section is present and populated regardless.
    keys = {section["key"] for section in report.content["sections"]}
    assert {"headline", "transactions", "exceptions", "extraction", "detail"} <= keys
    headline = next(s for s in report.content["sections"] if s["key"] == "headline")
    assert any(figure["value"] is not None for figure in headline["figures"])

    data = await storage_root.download(report.storage_ref)
    assert data.startswith(b"%PDF")


async def test_a_successful_ai_summary_is_labelled_as_the_model_s_work(
    db_session, storage_root, monkeypatch
):
    await seeded_period(db_session)
    user = await account(db_session, roles=[PlatformRole.APPROVER_HOD.value])
    await db_session.commit()

    from app.services.gemini_service import ExecutiveSummary

    async def answer(facts):
        # The model only ever sees already-computed figures, never a record.
        assert "percentage approved with no exception ever opened" in facts
        return ExecutiveSummary(summary="Throughput held steady across both streams.")

    monkeypatch.setattr(report_service, "summarize_reporting_period", answer)

    report = await report_service.generate(
        db_session, a_request(report_type="monthly"), requested_by=user, now=NOW
    )
    await db_session.commit()

    section = next(s for s in report.content["sections"] if s["kind"] == "ai_summary")
    assert section["text"] == "Throughput held steady across both streams."
    assert section["ai_generated"] is True
    assert report.ai_summary_error is None


# --- the schedule, riding 's sweep -----------------------------------------------------------------


async def test_the_daily_task_produces_the_report_for_the_day_that_has_ended(
    db_session, storage_root
):
    await seeded_period(db_session)

    result = await schedule.run_due(db_session, now=NOW)

    assert result.considered == 2
    assert len(result.generated) == 2

    daily = await db_session.scalar(select(Report).where(Report.report_type == "daily"))
    assert daily is not None
    # Nobody asked for it, so it is attributed to nobody rather than to an invented account.
    assert daily.generated_by_id is None
    assert daily.output_format == schedule.SCHEDULED_FORMAT
    expected = schedule.daily_due(NOW).period
    assert utc(daily.period_start) == expected.start
    assert utc(daily.period_end) == expected.end


async def test_the_monthly_task_covers_the_month_that_has_just_ended(db_session, storage_root):
    await schedule.run_due(db_session, now=NOW)

    monthly = await db_session.scalar(select(Report).where(Report.report_type == "monthly"))
    assert monthly is not None
    expected = schedule.monthly_due(NOW).period
    assert utc(monthly.period_start) == expected.start
    assert utc(monthly.period_end) == expected.end
    assert monthly.period_start.day == 1


async def test_the_schedule_is_idempotent_across_repeated_sweeps(db_session, storage_root):
    """The sweep runs every minute. It must produce one report per period, not one per tick."""
    first = await schedule.run_due(db_session, now=NOW)
    second = await schedule.run_due(db_session, now=NOW + timedelta(minutes=1))
    third = await schedule.run_due(db_session, now=NOW + timedelta(minutes=2))

    assert len(first.generated) == 2
    assert second.generated == []
    assert third.generated == []

    total = await db_session.scalar(select(func.count()).select_from(Report))
    assert total == 2


async def test_the_reused_sweep_loop_reaches_the_scheduled_tasks(db_session, monkeypatch):
    """'s periodic loop is the mechanism; this  only added tasks to it."""
    from app.services import integration_worker

    seen: list[str] = []

    async def spy(session, *, now=None):
        seen.append("asked")
        return schedule.ScheduleResult(considered=2)

    monkeypatch.setattr(schedule, "run_due", spy)
    await integration_worker.scheduled_reports_once()

    assert seen == ["asked"]
    # And there is no second worker module: the tasks are reached from the existing one.
    assert hasattr(integration_worker, "run_worker")
    assert integration_worker.report_schedule is schedule


def test_the_due_moments_come_from_configuration_rather_than_being_hard_coded():
    from app.core.config import settings

    due = schedule.daily_due(NOW)
    assert due.due_at.hour == settings.REPORT_DAILY_HOUR_UTC
    assert due.due_at.minute == settings.REPORT_DAILY_MINUTE_UTC
    assert due.due_at <= NOW

    monthly = schedule.monthly_due(NOW)
    assert monthly.due_at.day == settings.REPORT_MONTHLY_DAY
    assert monthly.due_at.hour == settings.REPORT_MONTHLY_HOUR_UTC
    assert monthly.due_at <= NOW


# --- template configuration ----------------------------------------------------------------------------


def test_every_shipped_template_names_only_sources_the_service_produces():
    for template in TEMPLATES:
        assert template.sections
        for section in template.sections:
            assert section.kind
            assert section.source
    assert template_for("daily").key == "daily_operations"
    assert template_for("monthly").wants_ai_summary is True
    assert template_for("adhoc").wants_ai_summary is False


def test_no_template_claims_the_file_itself_was_sent_anywhere():
    """The promise the disclosure has to keep, restated once distribution existed.

    It used to read "the platform has no outbound distribution capability", which stopped being
    true the day report distribution shipped. What is still true, and is what actually protects a
    reader holding a printed page, is narrower and more useful: *this file* is never sent by the
    platform. Where distribution is configured, recipients get a link and read the report in the
    platform. A document that reached somebody another way was put there by a person.
    """
    for template in TEMPLATES:
        joined = " ".join(template.disclosures).lower()
        assert "this file is never sent by the platform" in joined
        # And it must never claim the file was delivered to anybody.
        assert "emailed to" not in joined
        assert "attached" not in joined


async def test_a_section_naming_an_unknown_source_fails_the_build_rather_than_rendering_empty(
    db_session,
):
    from app.services.analytics.report_templates import ReportTemplate, SectionSpec

    broken = ReportTemplate(
        key="broken",
        title="Broken",
        report_type="adhoc",
        description="Names a source nothing produces.",
        sections=(SectionSpec(key="nope", title="Nope", kind="breakdown", source="not_a_source"),),
    )
    with pytest.raises(KeyError):
        report_service.build_content(
            broken,
            a_request(),
            {"headline": {}},
            reference="AGF-RPT-TEST",
            generated_by=None,
            generated_at=NOW,
        )


# --- validation ---------------------------------------------------------------------------------------------


def test_a_request_for_something_the_platform_does_not_produce_is_refused():
    from app.core.errors import BadRequestError

    with pytest.raises(BadRequestError):
        a_request(report_type="weekly")
    with pytest.raises(BadRequestError):
        a_request(output_format="csv")
    with pytest.raises(BadRequestError):
        a_request(stream="metals")
    with pytest.raises(BadRequestError):
        a_request(status="finished")
    with pytest.raises(BadRequestError):
        report_service.validate_request(
            report_type="adhoc",
            output_format="pdf",
            period=kpis.Period(start=NOW, end=NOW - timedelta(days=1)),
            stream="both",
            status_filter=None,
        )


# --- the prohibition this whole  turns on -------------------------------------------------


async def test_generating_a_report_alters_no_transaction_exception_approval_or_job(
    db_session, storage_root
):
    """The read-only guarantee, proved rather than asserted in a comment.

    Every row of every table this  reads is snapshotted, a report is generated over all of
    them, and the snapshot is compared. A single altered column anywhere fails this test.
    """
    from sqlalchemy import inspect

    from app.models.governance import ApprovalTask, ExceptionCase
    from app.models.integration import IntegrationJob
    from app.models.logistics import Shipment

    await seeded_period(db_session)
    user = await account(db_session, roles=[PlatformRole.ADMIN.value])
    await db_session.commit()

    watched = (TradeTransaction, ExceptionCase, ApprovalTask, IntegrationJob, Shipment)

    async def snapshot() -> dict:
        taken: dict = {}
        for model in watched:
            columns = [column.key for column in inspect(model).columns]
            for row in (await db_session.scalars(select(model))).all():
                taken[(model.__name__, str(row.id))] = tuple(
                    str(getattr(row, column)) for column in columns
                )
        return taken

    before = await snapshot()
    assert before, "the fixture must actually have written rows for this to prove anything"

    await report_service.generate(db_session, a_request(), requested_by=user, now=NOW)
    await db_session.commit()
    db_session.expire_all()

    assert await snapshot() == before


async def test_the_analytics_package_never_imports_a_write_path(db_session):
    """A structural check, so the guarantee survives somebody adding a convenient helper later."""
    import pathlib

    package = pathlib.Path(report_service.__file__).parent
    forbidden = (
        "transaction_service.submit",
        "approval_service.decide",
        "exception_service.resolve_case",
        "exception_service.open_case",
        "integration_service.dispatch",
        "shipment_service.apply_update",
    )
    for module in package.glob("*.py"):
        body = module.read_text()
        for name in forbidden:
            assert name not in body, f"{module.name} reaches a write path: {name}"
