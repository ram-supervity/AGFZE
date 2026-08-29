"""Turning assembled report content into real PDF and XLSX bytes.

Both renderers read the same content structure and neither of them knows what a section means.
They switch on the section's declared kind - a grid of figures, a breakdown, a table, a paragraph
- and on nothing else, which is what makes a template change a configuration change rather than a
code change here.

Two things are printed on every page and every sheet, deliberately and without a switch to turn
either off: the generation reference, so a printed page resolves back to the exact query behind
it, and the statement that the platform has not sent this document to anybody.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import pymupdf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.analytics.report_templates import (
    KIND_AI_SUMMARY,
    KIND_BREAKDOWN,
    KIND_KPI_GRID,
    KIND_NOTE,
    KIND_TABLE,
)

PDF_CONTENT_TYPE = "application/pdf"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

PAGE_WIDTH = 595.0
PAGE_HEIGHT = 842.0
MARGIN = 48.0
BODY_TOP = 92.0
BODY_BOTTOM = PAGE_HEIGHT - 58.0

FONT = "helv"
FONT_BOLD = "hebo"

INK = (0.09, 0.11, 0.15)
MUTED = (0.42, 0.45, 0.52)
RULE = (0.82, 0.84, 0.88)
ACCENT = (0.11, 0.31, 0.55)


# The base-14 PDF fonts are Latin-1, which has no em dash and no ellipsis. Using them anyway
# renders a stray dot in the middle of a table column, so the two places that need them use the
# ASCII forms instead.
EMPTY_CELL = "-"
ELLIPSIS = ".."


def _text(value: Any) -> str:
    if value is None:
        return EMPTY_CELL
    if isinstance(value, float):
        return f"{value:,.2f}".rstrip("0").rstrip(".") if value % 1 else f"{value:,.0f}"
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M UTC")
    return str(value)


UNIT_SUFFIX = {"percent": "%", "hours": "h"}


def _figure_text(figure: dict[str, Any]) -> str:
    """A figure with its unit attached. A duration printed as a bare number is not a duration."""
    if figure.get("value") is None:
        return EMPTY_CELL
    return f"{_text(figure['value'])}{UNIT_SUFFIX.get(str(figure.get('unit')), '')}"


def _moment(value: Any) -> str:
    if not value:
        return EMPTY_CELL
    try:
        return datetime.fromisoformat(str(value)).strftime("%Y-%m-%d %H:%M UTC")
    except ValueError:
        return str(value)


def _wrap(text: str, width: float, size: float, *, font: str = FONT) -> list[str]:
    """Greedy word wrap against the real glyph widths of the font being drawn with."""
    words = str(text).split()
    if not words:
        return [""]
    lines: list[str] = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        if pymupdf.get_text_length(candidate, fontname=font, fontsize=size) <= width:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def _truncate(text: str, width: float, size: float, *, font: str = FONT) -> str:
    rendered = str(text)
    if pymupdf.get_text_length(rendered, fontname=font, fontsize=size) <= width:
        return rendered
    while (
        rendered
        and pymupdf.get_text_length(rendered + ELLIPSIS, fontname=font, fontsize=size) > width
    ):
        rendered = rendered[:-1]
    return rendered + ELLIPSIS


class _Canvas:
    """A cursor down a stack of A4 pages, with the header and footer drawn once per page."""

    def __init__(self, content: dict[str, Any]) -> None:
        self.document = pymupdf.open()
        self.content = content
        self.page: Any = None
        self.y = 0.0
        self.width = PAGE_WIDTH - 2 * MARGIN
        self._new_page()

    def _new_page(self) -> None:
        self.page = self.document.new_page(width=PAGE_WIDTH, height=PAGE_HEIGHT)
        self.y = BODY_TOP
        self._header()
        self._footer()

    def _header(self) -> None:
        self.page.insert_text(
            (MARGIN, 44),
            "AGFZE Command Centre",
            fontname=FONT_BOLD,
            fontsize=11,
            color=ACCENT,
        )
        self.page.insert_text(
            (MARGIN, 60),
            str(self.content.get("title", "Report")),
            fontname=FONT,
            fontsize=9,
            color=MUTED,
        )
        # The reference, on every single page. A page torn out of a printed report still resolves.
        reference = str(self.content.get("generation_reference", ""))
        self.page.insert_text(
            (PAGE_WIDTH - MARGIN - pymupdf.get_text_length(reference, FONT_BOLD, 9), 44),
            reference,
            fontname=FONT_BOLD,
            fontsize=9,
            color=ACCENT,
        )
        self.page.draw_line(
            pymupdf.Point(MARGIN, 70), pymupdf.Point(PAGE_WIDTH - MARGIN, 70), color=RULE, width=0.6
        )

    def _footer(self) -> None:
        number = self.document.page_count
        self.page.draw_line(
            pymupdf.Point(MARGIN, BODY_BOTTOM + 14),
            pymupdf.Point(PAGE_WIDTH - MARGIN, BODY_BOTTOM + 14),
            color=RULE,
            width=0.6,
        )
        self.page.insert_text(
            (MARGIN, BODY_BOTTOM + 30),
            "Generated and stored in the platform. This file is never sent by the platform itself.",
            fontname=FONT,
            fontsize=7.5,
            color=MUTED,
        )
        label = f"Page {number}"
        self.page.insert_text(
            (PAGE_WIDTH - MARGIN - pymupdf.get_text_length(label, FONT, 7.5), BODY_BOTTOM + 30),
            label,
            fontname=FONT,
            fontsize=7.5,
            color=MUTED,
        )

    def space(self, height: float) -> None:
        if self.y + height > BODY_BOTTOM:
            self._new_page()

    def line(
        self,
        text: str,
        *,
        size: float = 9.5,
        font: str = FONT,
        colour: tuple[float, float, float] = INK,
        indent: float = 0.0,
        leading: float = 4.0,
    ) -> None:
        for row in _wrap(text, self.width - indent, size, font=font):
            self.space(size + leading)
            self.page.insert_text(
                (MARGIN + indent, self.y + size), row, fontname=font, fontsize=size, color=colour
            )
            self.y += size + leading

    def gap(self, height: float = 8.0) -> None:
        self.y = min(self.y + height, BODY_BOTTOM)

    def rule(self) -> None:
        self.space(10)
        self.page.draw_line(
            pymupdf.Point(MARGIN, self.y + 3),
            pymupdf.Point(PAGE_WIDTH - MARGIN, self.y + 3),
            color=RULE,
            width=0.6,
        )
        self.y += 10

    def table(self, columns: list[dict[str, Any]], rows: list[dict[str, Any]]) -> None:
        if not columns:
            return
        widths = _column_widths(columns, self.width)
        self._table_header(columns, widths)
        for index, row in enumerate(rows):
            if self.y + 16 > BODY_BOTTOM:
                self._new_page()
                self._table_header(columns, widths)
            if index % 2 == 1:
                self.page.draw_rect(
                    pymupdf.Rect(MARGIN, self.y - 1, PAGE_WIDTH - MARGIN, self.y + 13),
                    color=None,
                    fill=(0.965, 0.97, 0.98),
                )
            x = MARGIN
            for column, width in zip(columns, widths, strict=True):
                value = _truncate(_text(row.get(column["key"])), width - 5, 8.5)
                self.page.insert_text(
                    (x + 2, self.y + 9), value, fontname=FONT, fontsize=8.5, color=INK
                )
                x += width
            self.y += 14
        self.gap(6)

    def _table_header(self, columns: list[dict[str, Any]], widths: list[float]) -> None:
        self.space(20)
        x = MARGIN
        for column, width in zip(columns, widths, strict=True):
            # Bold and muted rather than uppercased: setting a header in capitals makes it
            # wider than the values under it, which is how a narrow column ends up with a
            # truncated heading over data that fits perfectly well.
            self.page.insert_text(
                (x + 2, self.y + 9),
                _truncate(str(column["label"]), width - 5, 7.5, font=FONT_BOLD),
                fontname=FONT_BOLD,
                fontsize=7.5,
                color=MUTED,
            )
            x += width
        self.y += 13
        self.page.draw_line(
            pymupdf.Point(MARGIN, self.y),
            pymupdf.Point(PAGE_WIDTH - MARGIN, self.y),
            color=RULE,
            width=0.6,
        )
        self.y += 3


def _column_widths(columns: list[dict[str, Any]], total: float) -> list[float]:
    weights = [float(column.get("weight", 1)) for column in columns]
    scale = total / sum(weights)
    return [weight * scale for weight in weights]


def render_pdf(content: dict[str, Any]) -> bytes:
    canvas = _Canvas(content)

    canvas.line(str(content.get("title", "Report")), size=16, font=FONT_BOLD)
    canvas.gap(2)
    canvas.line(_period_line(content), size=9, colour=MUTED)
    canvas.line(_provenance_line(content), size=9, colour=MUTED)
    canvas.gap(4)
    canvas.rule()

    for section in content.get("sections", []):
        _render_section(canvas, section)

    disclosures = content.get("disclosures") or []
    if disclosures:
        canvas.gap(6)
        canvas.rule()
        canvas.line("How to read this report", size=10, font=FONT_BOLD)
        for note in disclosures:
            canvas.line(f"-  {note}", size=8.5, colour=MUTED, indent=4)

    canvas.gap(6)
    canvas.line(
        f"Generation reference {content.get('generation_reference')} — quote it to resolve this "
        "document back to the exact query, parameters and audit record behind it.",
        size=8.5,
        colour=ACCENT,
    )

    buffer = canvas.document.tobytes()
    canvas.document.close()
    return buffer


def _render_section(canvas: _Canvas, section: dict[str, Any]) -> None:
    canvas.gap(6)
    canvas.line(str(section.get("title", "")), size=11.5, font=FONT_BOLD)
    if section.get("description"):
        canvas.line(str(section["description"]), size=8.5, colour=MUTED)
    canvas.gap(2)

    kind = section.get("kind")
    if kind == KIND_KPI_GRID:
        for figure in section.get("figures", []):
            canvas.line(f"{figure.get('label')}:  {_figure_text(figure)}", size=9.5, indent=4)
            if figure.get("note"):
                canvas.line(str(figure["note"]), size=8, colour=MUTED, indent=14, leading=3)
    elif kind in (KIND_BREAKDOWN, KIND_TABLE):
        rows = section.get("rows", [])
        if rows:
            canvas.table(list(section.get("columns", [])), rows)
        else:
            canvas.line("Nothing in this period.", size=9, colour=MUTED, indent=4)
    elif kind == KIND_AI_SUMMARY:
        if section.get("text"):
            canvas.line("AI-generated summary", size=8, font=FONT_BOLD, colour=MUTED, indent=4)
            canvas.line(str(section["text"]), size=9.5, indent=4)
        else:
            canvas.line(
                str(
                    section.get("unavailable_reason")
                    or "The AI summary is unavailable for this report."
                ),
                size=9,
                colour=MUTED,
                indent=4,
            )
    elif kind == KIND_NOTE:
        canvas.line(str(section.get("text", "")), size=9.5, indent=4)
    canvas.gap(4)


def _period_line(content: dict[str, Any]) -> str:
    period = content.get("period") or {}
    stream = content.get("stream", "both")
    status = content.get("status_filter")
    parts = [
        f"Period {_moment(period.get('start'))} to {_moment(period.get('end'))}",
        f"stream: {'both streams' if stream == 'both' else stream}",
    ]
    parts.append(f"status: {status}" if status else "status: every status")
    return "  ·  ".join(parts)


def _provenance_line(content: dict[str, Any]) -> str:
    who = content.get("generated_by") or "the platform (scheduled generation)"
    return f"Generated {_moment(content.get('generated_at'))} by {who}"


# --- XLSX ----------------------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1C4E8C")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13)
MUTED_FONT = Font(color="6B7280", size=9)


def render_xlsx(content: dict[str, Any]) -> bytes:
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Report"
    _write_cover(cover, content)

    used: set[str] = {cover.title}
    for section in content.get("sections", []):
        kind = section.get("kind")
        if kind not in (KIND_BREAKDOWN, KIND_TABLE):
            continue
        sheet = workbook.create_sheet(_sheet_name(str(section.get("title", "Section")), used))
        _write_table_sheet(sheet, section, content)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _sheet_name(title: str, used: set[str]) -> str:
    cleaned = "".join(character for character in title if character not in "[]:*?/\\")[:31].strip()
    candidate = cleaned or "Section"
    suffix = 2
    while candidate in used:
        candidate = f"{cleaned[:28]} {suffix}"
        suffix += 1
    used.add(candidate)
    return candidate


def _write_cover(sheet: Any, content: dict[str, Any]) -> None:
    period = content.get("period") or {}
    sheet["A1"] = str(content.get("title", "Report"))
    sheet["A1"].font = TITLE_FONT

    meta = [
        ("Generation reference", content.get("generation_reference")),
        ("Generated at", _moment(content.get("generated_at"))),
        ("Generated by", content.get("generated_by") or "the platform (scheduled generation)"),
        ("Report type", content.get("report_type")),
        ("Template", content.get("template_key")),
        ("Period start", _moment(period.get("start"))),
        ("Period end", _moment(period.get("end"))),
        ("Business stream", content.get("stream")),
        ("Status filter", content.get("status_filter") or "every status"),
    ]
    row = 3
    for label, value in meta:
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        sheet.cell(row=row, column=2, value=_text(value))
        row += 1

    row += 1
    for section in content.get("sections", []):
        if section.get("kind") == KIND_KPI_GRID:
            sheet.cell(row=row, column=1, value=str(section.get("title", ""))).font = Font(
                bold=True, size=11
            )
            row += 1
            for figure in section.get("figures", []):
                sheet.cell(row=row, column=1, value=str(figure.get("label")))
                sheet.cell(row=row, column=2, value=figure.get("value"))
                sheet.cell(row=row, column=3, value=str(figure.get("unit", "")))
                if figure.get("note"):
                    note = sheet.cell(row=row, column=4, value=str(figure["note"]))
                    note.font = MUTED_FONT
                row += 1
            row += 1
        elif section.get("kind") == KIND_AI_SUMMARY:
            sheet.cell(row=row, column=1, value=str(section.get("title", ""))).font = Font(
                bold=True, size=11
            )
            row += 1
            body = section.get("text") or section.get("unavailable_reason") or ""
            cell = sheet.cell(row=row, column=1, value=str(body))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if section.get("text"):
                sheet.cell(row=row, column=2, value="AI-generated").font = MUTED_FONT
            row += 2

    for note in content.get("disclosures") or []:
        sheet.cell(row=row, column=1, value=str(note)).font = MUTED_FONT
        row += 1

    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 14
    sheet.column_dimensions["D"].width = 70
    sheet.freeze_panes = "A3"


def _write_table_sheet(sheet: Any, section: dict[str, Any], content: dict[str, Any]) -> None:
    columns = list(section.get("columns", []))
    rows = list(section.get("rows", []))

    sheet.cell(row=1, column=1, value=str(section.get("title", ""))).font = TITLE_FONT
    sheet.cell(
        row=2, column=1, value=f"Generation reference {content.get('generation_reference')}"
    ).font = MUTED_FONT
    if section.get("description"):
        sheet.cell(row=3, column=1, value=str(section["description"])).font = MUTED_FONT

    header_row = 5
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=str(column["label"]))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        sheet.column_dimensions[get_column_letter(index)].width = int(column.get("width", 20))

    for offset, row in enumerate(rows, start=header_row + 1):
        for index, column in enumerate(columns, start=1):
            sheet.cell(row=offset, column=index, value=_cell_value(row.get(column["key"])))

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)


def _cell_value(value: Any) -> Any:
    """Numbers stay numbers so the recipient can sum a column; everything else becomes text."""
    if value is None:
        return None
    if isinstance(value, int | float | bool):
        return value
    return str(value)
